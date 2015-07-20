# -*- coding: utf-8 -*-
import json
# from datetime import date

from django.http import JsonResponse, HttpResponse
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.core.urlresolvers import reverse
from django.contrib import messages
from django.utils.translation import ugettext as _

from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.styles import Style, Font, Alignment
from openpyxl.worksheet.dimensions import ColumnDimension, RowDimension
from openpyxl.cell import get_column_letter

import nepdate

from core.models import app_setting, FiscalYear
from app.utils.helpers import invalid, save_model, empty_to_none
from users.models import group_required

from inventory.filters import InventoryItemFilter

from inventory.forms import ItemForm, CategoryForm, DemandForm, PurchaseOrderForm, HandoverForm, EntryReportForm, \
    ItemLocationForm, DepreciationForm, ItemInstanceForm

from inventory.models import Depreciation, Demand, ItemInstance, \
    DemandRow, delete_rows, Item, Category, PurchaseOrder, PurchaseOrderRow, \
    InventoryAccount, Handover, HandoverRow, EntryReport, EntryReportRow, set_transactions, JournalEntry, \
    InventoryAccountRow, Transaction, Inspection, InspectionRow, YearlyReport, YearlyReportRow, ItemLocation, Release

from inventory.serializers import DepreciationSerializer, DemandSerializer, ItemSerializer, PurchaseOrderSerializer, \
    HandoverSerializer, EntryReportSerializer, EntryReportRowSerializer, InventoryAccountRowSerializer, \
    TransactionSerializer, ItemLocationSerializer, ItemInstanceSerializer


def list_transactions(request):
    if request.POST:
        param = request.POST
        date_list = request.POST.getlist('myDate')
        start_date_bs = date_list[0]
        end_date_bs = date_list[1]
        if nepdate.is_valid(start_date_bs) and nepdate.is_valid(end_date_bs):
            start_date_ad = nepdate.string_from_tuple(nepdate.bs2ad(start_date_bs))
            end_date_ad = nepdate.string_from_tuple(nepdate.bs2ad(end_date_bs))
            obj = Transaction.objects.filter(journal_entry__date__range=[start_date_ad, end_date_ad])
            return render(request, "transaction_list.html", {'objects': obj})
        elif start_date_bs == '' and end_date_bs == '':
            messages.error(request, "Enter Date")
        else:
            messages.error(request, "Invalid Date")
    obj = Transaction.objects.all()
    return render(request, "transaction_list.html", {'objects': obj})


def xlsx_formula(ws, start_row, start_column, end_row, end_column, value):
    first_cell_column_id = str(ws.cell(row=start_row, column=start_column).column)
    first_cell_row_id = str(ws.cell(row=start_row, column=start_column).row)
    second_cell_column_id = str(ws.cell(row=end_row, column=end_column).column)
    second_cell_row_id = str(ws.cell(row=end_row, column=end_column).row)
    if value == "=SUM":
        return "=SUM(" + first_cell_column_id + first_cell_row_id + ":" + second_cell_column_id + second_cell_row_id + ")"
    if value == "=PRODUCT":
        return "=PRODUCT(" + first_cell_column_id + first_cell_row_id + "," + second_cell_column_id + second_cell_row_id + ")"
    return ''


def insert_row(ws, row, column, args, extra_col_value1=None, extra_col_value2=None):
    for i, value in enumerate(args):
        cell = ws.cell(row=row, column=i + 1)
        if value == "=PRODUCT":
            quantity = ws.cell(row=row, column=extra_col_value1).value
            rate = ws.cell(row=row, column=extra_col_value2).value
            cell.value = "=PRODUCT(" + str(quantity) + "," + str(rate) + ")"
        elif value == "VAT":
            rate = ws.cell(row=row, column=extra_col_value2).value
            # vat = rate * .13
            # cell.value = vat
            cell.value = "=PRODUCT(" + str(rate) + "," + ".13 )"
        elif value == "UnitPrice":
            rate = ws.cell(row=row, column=extra_col_value2).value
            vat = ws.cell(row=row, column=extra_col_value2 + 1)
            cell.value = xlsx_formula(ws, row, extra_col_value2, row, extra_col_value2 + 1, "=SUM")
            # vat = rate * .13
            # unit_price = rate + vat
            # cell.value = unit_price
        elif value == "Total":
            quantity = ws.cell(row=row, column=extra_col_value1).value
            rate = ws.cell(row=row, column=extra_col_value2).value
            vat = rate * .13
            unit_price = rate + vat
            total = quantity * unit_price
            other_expenses = ws.cell(row=row, column=extra_col_value2 + 3).value
            grand_total = total + other_expenses
            cell.value = grand_total
        else:
            cell.value = value
    return row + 1


def merge_and_add(ws, start_row, start_column, end_row, end_column, value):
    ws.merge_cells(start_row=start_row, start_column=start_column, end_row=end_row, end_column=end_column)
    cell = ws.cell(row=start_row, column=start_column)
    cell.value = value
    return cell


def add_cell_value(ws, row, column, value):
    cell = ws.cell(row=row, column=column)
    cell.value = value
    return cell


def convert_demand(request, id):
    demand = get_object_or_404(Demand, id=id)
    wb = Workbook()
    ws = wb.active
    ws.merge_cells('A2:H2')
    header = ws.cell('A2')
    header.value = app_setting.header_for_forms
    header.style = Style(
        font=Font(
            bold=True,
            size=24),
        alignment=Alignment(
            horizontal='center'),
    )
    ws.merge_cells('A4:H4')
    report_name = ws.cell('A4')
    report_name.value = _("Demand Form")
    report_name.style = Style(
        font=Font(
            # bold=True,
            size=18),
        alignment=Alignment(
            horizontal='center'),
    )
    release_no = ws.cell('F5')
    release_no.value = _("Release No.") + _(str(demand.release_no))
    release_no = ws.cell('H5')
    release_no.value = _("Fiscal Year") + ":- " + _(str(demand.fiscal_year))

    ws['A5'] = "श्री प्रमुख,"
    ws['A6'] = "भण्डार शाखा"

    table_header = [_('SN'), _('Item Name'), _('Specification'), _('Item Quantity'), \
                    _('Unit'), _('Released Item Quantity'), _('Inventory Account No.'), _('Remarks')]
    row_index = insert_row(ws, 7, 1, table_header)
    for row in demand.rows.all():
        data = [row.sn, row.item.name, row.specification, row.quantity, row.unit, row.release_quantity,
                row.item.property_classification_reference_number, row.remarks]
        row_index = insert_row(ws, row_index, 1, data)
    ws.column_dimensions[get_column_letter(1)].width = 4
    ws.column_dimensions[get_column_letter(2)].width = 40
    ws.column_dimensions[get_column_letter(3)].width = 25
    ws.column_dimensions[get_column_letter(4)].width = 15
    ws.column_dimensions[get_column_letter(6)].width = 20
    ws.column_dimensions[get_column_letter(7)].width = 20
    ws.column_dimensions[get_column_letter(8)].width = 20

    ColumnDimension(ws, index="B", width=200, customWidth=True)
    RowDimension(ws, index=2, ht=300, customHeight=True)
    footer_base = row_index + 1
    ws.cell(row=footer_base, column=2).value = _("Demandee's Signature") + ":- "
    ws.cell(row=footer_base, column=6).value = _("(a)") + _("Buy from market")
    ws.cell(row=footer_base + 2, column=2).value = _("Name") + ":- " + _(demand.demandee.username)
    ws.cell(row=footer_base + 1, column=6).value = _("(b)") + _("Lend from store")
    ws.cell(row=footer_base + 3, column=2).value = _("Date") + ":- " + _(str(demand.date))
    ws.cell(row=footer_base + 3, column=6).value = _("Signature of the orderer") + ":- "
    ws.cell(row=footer_base + 4, column=2).value = _("Purpose") + ":- " + demand.purpose
    ws.cell(row=footer_base + 4, column=6).value = _("Date") + ":- "
    ws.cell(row=footer_base + 5, column=2).value = _("Signature of the one who enters in Inventory Account") + ":- "
    ws.cell(row=footer_base + 5, column=6).value = _("Signature of the receiver") + ":- "
    ws.cell(row=footer_base + 6, column=2).value = _("Date") + ":- "
    ws.cell(row=footer_base + 6, column=6).value = _("Date") + ":- "

    response = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="demand-report.xlsx"'
    return response


def convert_purchase_order(request, id):
    purchase_order = get_object_or_404(PurchaseOrder, id=id)
    wb = Workbook()
    ws = wb.active

    # Header
    header = merge_and_add(ws, 2, 1, 2, 9, app_setting.header_for_forms)
    header.style = Style(
        font=Font(
            bold=True,
            size=24),
        alignment=Alignment(
            horizontal='center'),
    )
    report_name = merge_and_add(ws, 4, 1, 4, 9, "Purchase Order")
    report_name.style = Style(
        font=Font(
            # bold=True,
            size=18),
        alignment=Alignment(
            horizontal='center'),
    )
    ws.cell('A7').value = _("Shree") + ":- " + _(purchase_order.party.name)
    ws.cell('H7').value = _("Purchase Order") + _("No.") + ":- " + _(str(purchase_order.order_no))
    ws.cell('A8').value = _("Address") + ":- " + _(purchase_order.party.address)
    ws.cell('H8').value = _("Date") + ":- " + _(str(purchase_order.date))
    ws.cell('A10').value = _("VAT/PAN") + ":- " + _(str(purchase_order.party.pan_no))
    merge_and_add(ws, 10, 1, 10, 2, "देहाय बमोजिमका सामानहरु")
    ws.cell(row=10, column=3).value = purchase_order.due_days
    merge_and_add(ws, 10, 4, 10, 9, "दिन भित्र यस कार्यालयमा दाखिला गरि विल / इन्भ्वाईस प्रस्तुत गर्नु होला ।")
    ws.merge_cells('A13:A14');
    ws.merge_cells('B13:B14');
    ws.merge_cells('C13:C14');
    ws.merge_cells('D13:D14');
    ws.merge_cells('E13:E14');
    ws.merge_cells('F13:F14');
    ws.merge_cells('I13:I14');
    ws.merge_cells('G13:H13')

    # Table Head
    table_header = [_('SN'), _('Budget Title No.'), _('Particular'), _('Specification'), _('Item Quantity'), \
                    _('Unit'), _("Price")]
    row_index = insert_row(ws, 13, 1, table_header)
    ws.cell('I13').value = _("Remarks")
    table_sub_header = [_("Rate"), _("Total Amount")]
    ws.cell('G14').value = _("Rate")
    ws.cell('H14').value = _("Total Amount")
    row_index = row_index + 1

    # Table body
    for row in purchase_order.rows.all().order_by("sn"):
        data = [row.sn, row.budget_title_no, row.item.name, row.specification, row.quantity, \
                row.unit, row.rate, "=PRODUCT", row.remarks]
        row_index = insert_row(ws, row_index, 1, data, 5, 7)
    last_row_number = row_index - 1
    total = merge_and_add(ws, row_index, 1, row_index, 6, _("Total"))
    ws.cell(row=row_index, column=8).value = xlsx_formula(ws, 15, 8, last_row_number, 8, "=SUM")
    vat = merge_and_add(ws, row_index + 1, 1, row_index + 1, 6, _("13% VAT"))
    ws.cell(row=row_index + 1, column=8).value = "=PRODUCT(" + str(ws.cell(row=row_index, column=8).column) + str(
        ws.cell(row=row_index, column=8).row) + "," + ".13)"
    grand_total = merge_and_add(ws, row_index + 2, 1, row_index + 2, 6, _("Grand Total"))
    ws.cell(row=row_index + 2, column=8).value = xlsx_formula(ws, row_index, 8, row_index + 1, 8, "=SUM")

    # Footer
    add_cell_value(ws, row_index + 4, 2, _("Faantwaala's") + _('Signature'))
    add_cell_value(ws, row_index + 4, 7, _('Section') + _("Head's") + _('Signature'))
    add_cell_value(ws, row_index + 5, 2, _("Date"))
    add_cell_value(ws, row_index + 5, 7, _("Date"))
    add_cell_value(ws, row_index + 5, 7, _("Date"))
    fill_by_admin = add_cell_value(ws, row_index + 7, 1, _("To be filled by financial administration section") + ":- ")
    fill_by_admin.font = Font(bold=True)
    add_cell_value(ws, row_index + 9, 1,
                   "माथि उल्लेखि सामानहरु बजट उपशिर्षक न. .............. को खर्च शिर्षक न. .......... बाट भुक्तानी दिन बजेट बाँकी देखिन्छ / देखिदैंन ।")
    add_cell_value(ws, row_index + 10, 7, _('Accounting') + _("Head's") + _('Signature') + ":- ")
    add_cell_value(ws, row_index + 11, 7, _("Date"))
    add_cell_value(ws, row_index + 13, 7, _("Signature of Head of Office") + ":- ")
    add_cell_value(ws, row_index + 14, 7, _("Date"))
    add_cell_value(ws, row_index + 15, 1,
                   "माथि उल्लेखित सामानहरु मिति .......................... भित्र................................... कार्यालयमा बुझाउने छु भनी सहिछाप गर्ने ।")
    add_cell_value(ws, row_index + 17, 3, _("Firm's Name"))
    merge_and_add(ws, row_index + 17, 5, row_index + 17, 6, _("Signature"))
    add_cell_value(ws, row_index + 17, 8, _("Firm's Name"))
    response = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="purchase-order.xlsx"'
    return response


def convert_entry_report(request, id):
    entry_report = get_object_or_404(EntryReport, id=id)
    wb = Workbook()
    ws = wb.active
    row_index = 9
    # Header
    header = merge_and_add(ws, 2, 1, 2, 13, app_setting.header_for_forms)
    header.style = Style(
        font=Font(
            bold=True,
            size=24),
        alignment=Alignment(
            horizontal='center'),
    )
    report_name = merge_and_add(ws, 4, 1, 4, 13, _("Entry Report"))
    report_name.style = Style(
        font=Font(
            # bold=True,
            size=18),
        alignment=Alignment(
            horizontal='center'),
    )
    add_cell_value(ws, 6, 1, _("Entry Report") + ' ' + _("No." + ":- ") + _(str(entry_report.entry_report_no)))

    # Table Head
    table_header = [_('SN'), _('Inventory Account Page No.'), _('Inventory Classification Reference No.'),
                    _("Item's Name"), _('Specification'), \
                    _('Unit'), _('Quantity')]
    for i, value in enumerate(table_header):
        merge_and_add(ws, 7, i + 1, 8, i + 1, value)
    merge_and_add(ws, 7, 8, 7, 12, _('Price') + "(" + _('As per Invoice') + ")")
    table_sub_header = [_('Rate per Unit'), _('VAT') + ' ' + _("per") + ' ' + _("Unit"), _("Unit") + ' ' + _("Price"), \
                        _('Other') + ' ' + _('Expenses'), _("Total")]
    for i, value in enumerate(table_sub_header):
        add_cell_value(ws, 8, i + 8, value)
    merge_and_add(ws, 7, 13, 8, 13, _("Remarks"))

    # Table body
    for row in entry_report.rows.all().order_by("sn"):
        data = [row.sn, row.item.account.account_no, row.item.property_classification_reference_number, row.item.name,
                row.specification, row.unit, \
                row.quantity, row.rate, "VAT", "UnitPrice", row.other_expenses, "Total", row.remarks]
        row_index = insert_row(ws, row_index, 1, data, 7, 8)

    # Footer
    merge_and_add(ws, row_index + 1, 1, row_index + 1, 6,
                  "माथि उल्लेखित सामानहरु खरिद आदेश नम्बर/हस्तान्तरण फारम नम्बर")
    add_cell_value(ws, row_index + 1, 7, _(str(entry_report.source.order_no)))
    add_cell_value(ws, row_index + 1, 8, "मिति")
    add_cell_value(ws, row_index + 1, 9, _(str(entry_report.source.date)))
    add_cell_value(ws, row_index + 1, 10, 'अनुसार श्री')
    add_cell_value(ws, row_index + 1, 11, _(str(entry_report.source.party)))
    add_cell_value(ws, row_index + 1, 12, "बाट प्राप्त हुन आएको हुँदा जाँची गन्ती गरी हेर्दा ठीक दुरुस्त भएकोले")
    add_cell_value(ws, row_index + 2, 1, "खातामा आम्दानी बाँधेको प्रमाणित गर्दछु ।")
    add_cell_value(ws, row_index + 4, 1, _("Faantwaala's") + _('Signature'))
    add_cell_value(ws, row_index + 5, 1, _("Name"))
    add_cell_value(ws, row_index + 6, 1, _("Designation"))
    add_cell_value(ws, row_index + 7, 1, _("Date"))
    add_cell_value(ws, row_index + 4, 6, _("Signature of Verifying Section Head") + ":- ")
    add_cell_value(ws, row_index + 5, 6, _("Name"))
    add_cell_value(ws, row_index + 6, 6, _("Designation"))
    add_cell_value(ws, row_index + 7, 6, _("Date"))
    add_cell_value(ws, row_index + 4, 10, _("Signature of Head of Office") + ":- ")
    add_cell_value(ws, row_index + 5, 10, _("Name"))
    add_cell_value(ws, row_index + 6, 10, _("Designation"))
    add_cell_value(ws, row_index + 7, 10, _("Date"))

    response = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="entry-report.xlsx"'
    return response


def remove_transaction_duplicate(object):
    compare_name = []
    compare_rate = []
    object_list = []
    for o in object:
        if o.account.name not in compare_name:
            compare_name.append(o.account.name)
            compare_rate.append(o.journal_entry.creator.rate)
            object_list.append(o)
        if o.account.name in compare_name and o.journal_entry.creator.rate not in compare_rate:
            compare_rate.append(o.journal_entry.creator.rate)
            object_list.append(o)
    compare_name = []
    return object_list


def remove_transaction_duplicate_for_yearly_report(object):
    compare_name = []
    object_list = []
    for o in object:
        if o.account.name not in compare_name:
            compare_name.append(o.account.name)
            object_list.append(o)
    compare_name = []
    return object_list


def yearly_report(request):
    obj = Transaction.objects.filter(cr_amount=None)
    transaction_without_duplication = remove_transaction_duplicate_for_yearly_report(obj)
    data = TransactionSerializer(transaction_without_duplication, many=True).data
    return render(request, 'yearly_report.html', {'data': data})


def yearly_report_list(request):
    obj = YearlyReport.objects.all()
    return render(request, 'yearly_report_list.html', {'obj': obj})


def yearly_report_detail(request, id):
    obj = YearlyReport.objects.get(pk=id)
    rows = obj.rows.order_by("sn")
    return render(request, 'yearly_report_detail.html', {'obj': obj, 'rows': rows})


@group_required('Store Keeper', 'Chief')
def delete_yearly_report(request, id):
    obj = get_object_or_404(YearlyReport, id=id)
    obj.delete()
    return redirect(reverse('yearly_report_list'))


def save_yearly_report(request):
    if request.is_ajax():
        params = json.loads(request.body)
    dct = {'rows': {}}
    if params.get('release_no') == '':
        params['release_no'] = None
    object_values = {'release_no': params.get('release_no'), 'fiscal_year': FiscalYear.get(app_setting.fiscal_year)}
    if params.get('id'):
        obj = YearlyReport.objects.get(id=params.get('id'))
    else:
        obj = YearlyReport()
    try:
        obj = save_model(obj, object_values)
        dct['id'] = obj.id
        model = YearlyReportRow
        for index, row in enumerate(params.get('table_view').get('rows')):
            values = {'sn': index + 1, 'account_no': row.get('account_no'),
                      'property_classification_reference_number': row.get('inventory_classification_reference_no'),
                      'item_name': row.get('item_name'), 'income': row.get('total_dr_amount'),
                      'expense': row.get('expense'), 'remaining': row.get('current_balance'),
                      'remarks': row.get('remarks'), 'yearly_report': obj}
            submodel, created = model.objects.get_or_create(id=row.get('id'), defaults=values)
            if not created:
                submodel = save_model(submodel, values)
            dct['rows'][index] = submodel.id
        delete_rows(params.get('table_view').get('deleted_rows'), model)
    except Exception as e:
        if hasattr(e, 'messages'):
            dct['error_message'] = '; '.join(e.messages)
        elif str(e) != '':
            dct['error_message'] = str(e)
        else:
            dct['error_message'] = 'Error in form data!'
    return JsonResponse(dct)


def inspection_report_list(request):
    obj = Inspection.objects.all()
    return render(request, 'inspection_list.html', {'obj': obj})


def inspection_report_detail(request, id):
    obj = Inspection.objects.get(pk=id)
    rows = obj.rows.order_by("sn")
    return render(request, 'inspection_detail.html', {'obj': obj, 'rows': rows})


def inspection_report(request):
    obj = Transaction.objects.filter(cr_amount=None)
    transaction_without_duplication = remove_transaction_duplicate(obj)
    data = TransactionSerializer(transaction_without_duplication, many=True).data
    return render(request, 'inspection_report.html', {'obj': transaction_without_duplication, 'data': data})


@group_required('Store Keeper', 'Chief')
def delete_inspection_report(request, id):
    obj = get_object_or_404(Inspection, id=id)
    obj.delete()
    return redirect(reverse('inspection_report_list'))


def save_inspection_report(request):
    if request.is_ajax():
        params = json.loads(request.body)
    dct = {'rows': {}}
    if params.get('release_no') == '':
        params['release_no'] = None
    object_values = {'release_no': params.get('release_no'), 'fiscal_year': FiscalYear.get(app_setting.fiscal_year)}
    if params.get('id'):
        obj = Inspection.objects.get(id=params.get('id'))
    else:
        obj = Inspection()
    try:
        obj = save_model(obj, object_values)
        dct['id'] = obj.id
        model = InspectionRow
        for index, row in enumerate(params.get('table_view').get('rows')):
            values = {'sn': index + 1, 'account_no': row.get('account_no'),
                      'property_classification_reference_number': row.get('inventory_classification_reference_no'),
                      'item_name': row.get('item_name'), 'unit': row.get('unit'),
                      'quantity': row.get('total_dr_amount'), 'rate': row.get('rate'), 'price': row.get('price'),
                      'matched_number': empty_to_none(row.get('match_number')),
                      'unmatched_number': empty_to_none(row.get('unmatch_number')),
                      'decrement': empty_to_none(row.get('decrement')),
                      'increment': empty_to_none(row.get('increment')),
                      'decrement_increment_price': empty_to_none(row.get('decrement_increment_price')),
                      'good': empty_to_none(row.get('good')), 'bad': empty_to_none(row.get('bad')),
                      'remarks': row.get('remarks'), 'inspection': obj}

            submodel, created = model.objects.get_or_create(id=row.get('id'), defaults=values)
            if not created:
                submodel = save_model(submodel, values)
            dct['rows'][index] = submodel.id
        delete_rows(params.get('table_view').get('deleted_rows'), model)
    except Exception as e:
        if hasattr(e, 'messages'):
            dct['error_message'] = '; '.join(e.messages)
        elif str(e) != '':
            dct['error_message'] = str(e)
        else:
            dct['error_message'] = 'Error in form data!'
    return JsonResponse(dct)


def depreciation_report(request):
    obj = Transaction.objects.filter(account__item__depreciation__depreciate_value__gte=0, cr_amount=None)
    transaction_without_duplication = remove_transaction_duplicate(obj)
    # depreciate_object_list = []
    # for depreciate in transaction_without_duplication:
    #     depreciate_object_list.append(depreciate.account.item.depreciation)
    transaction = TransactionSerializer(transaction_without_duplication, many=True).data
    # depreciate_item = DepreciationSerializer(depreciate_object_list, many=True).data
    # depreciate_object_list = []
    # import pdb; pdb.set_trace()
    return render(request, "depreciation_report.html", {'data': transaction})


@login_required
def item_form(request, id=None):
    if id:
        item = get_object_or_404(Item, id=id)
        scenario = 'Update'
        depreciation_data = DepreciationSerializer(item.depreciation).data
        item_instances = ItemInstance.objects.filter(item__id=item.id)
    else:
        item = Item()
        scenario = 'Create'
        depreciation = Depreciation(depreciate_type="Fixed percentage", depreciate_value=0, time=0, time_type='Year(s)')
        depreciation_data = DepreciationSerializer(depreciation).data
        item_instances = []
    if request.POST:
        form = ItemForm(data=request.POST, instance=item, user=request.user)
        if form.is_valid():
            item = form.save(commit=False)
            property_name = request.POST.getlist('property_name')
            item_property = request.POST.getlist('property')
            time = request.POST.get('time')
            depreciate_value = request.POST.get('depreciate_value')
            depreciate_type = request.POST.get('depreciate_type')
            time_type = request.POST.get('time_type')
            depreciation_id = request.POST.get('depreciation_id')
            if depreciation_id == '':
                dep = Depreciation(time=time, depreciate_value=depreciate_value, depreciate_type=depreciate_type,
                                   time_type=time_type)
                dep.save()
            else:
                dep = Depreciation.objects.get(pk=depreciation_id)
                dep.time = time
                dep.depreciate_value = depreciate_value
                dep.depreciate_type = depreciate_type
                dep.time_type = time_type
                dep.save()
            other_properties = {}
            for key, value in zip(property_name, item_property):
                other_properties[key] = value
            # other_properties_json = json.dumps(other_properties, sort_keys=True, indent=4)
            item.other_properties = other_properties
            item.depreciation = dep
            item.save(account_no=form.cleaned_data['account_no'], opening_balance=form.cleaned_data['opening_balance'])
            if request.is_ajax():
                return render(request, 'callback.html', {'obj': ItemSerializer(item).data})
            return redirect('/inventory/items/')
    else:
        form = ItemForm(instance=item, user=request.user)
        depreciation_form = DepreciationForm()
    if request.is_ajax():
        base_template = 'modal.html'
    else:
        base_template = 'inventory_base.html'
    return render(request, 'item_form.html', {
        'scenario': scenario,
        'form': form,
        'base_template': base_template,
        'item_data': item.other_properties,
        'depreciation_form': depreciation_form,
        'depreciation_data': depreciation_data,
        'item_instances': item_instances,
    })


def item_instance_form(request, id):
    item = get_object_or_404(ItemInstance, id=id)
    if request.POST:
        form = ItemInstanceForm(data=request.POST, instance=item)
        if form.is_valid():
            item_instance = form.save(commit=False)
            property_name = request.POST.getlist('property_name')
            item_property = request.POST.getlist('property')
            other_properties = {}
            for key, value in zip(property_name, item_property):
                other_properties[key] = value
            item_instance.other_properties = other_properties
            item_instance.save()
            return redirect(reverse('update_inventory_item', kwargs={'id': item_instance.item_id}))
    else:
        form = ItemInstanceForm(instance=item)
    return render(request, 'item_instance_form.html', {'form': form, 'item_data': item.other_properties})


@group_required('Store Keeper', 'Chief')
def delete_inventory_item(request, id):
    obj = get_object_or_404(Item, id=id)
    obj.delete()
    return redirect('/inventory/items/')


@group_required('Store Keeper', 'Chief')
def list_inventory_items(request):
    objects = Item.objects.all()
    filtered_items = InventoryItemFilter(request.GET, queryset=objects)
    return render(request, 'list_inventory_items.html', {'objects': filtered_items})


@login_required
def list_demand_forms(request):
    if request.user.in_group('Store Keeper') or request.user.in_group('Chief'):
        objects = Demand.objects.all()
    else:
        objects = Demand.objects.filter(demandee=request.user)
    return render(request, 'list_demand_forms.html', {'objects': objects})


@login_required
def delete_demand(request, id):
    if request.user.in_group('Store Keeper') or request.user.in_group('Chief'):
        obj = get_object_or_404(Demand, id=id)
    else:
        obj = get_object_or_404(Demand, id=id, demandee=request.user)
    obj.delete()
    return redirect(reverse('list_demand_forms'))


@login_required
def items_as_json(request):
    items = Item.objects.all()
    items_data = ItemSerializer(items, many=True).data
    return JsonResponse(items_data, safe=False)


@login_required
def item_instances_as_json(request):
    item_instances = ItemInstance.objects.filter(location__name='Store')
    instances = {}
    for instance in item_instances:
        if not instance.item_id in instances.keys():
            instances[instance.item_id] = {}
        s = str(instance.item_rate)
        rate = s.rstrip('0').rstrip('.') if '.' in s else s
        instance.other_properties['rate'] = rate
        # property = cPickle.dumps(item.other_properties)
        prop = json.dumps(instance.other_properties).replace(' ', '')
        if not prop in instances[instance.item_id].keys():
            instances[instance.item_id][prop] = []
        instances[instance.item_id][prop].append(instance.id)
    lst = []
    for key, value in instances.iteritems():
        instance_lst = []
        for pro, instance_ids in value.iteritems():
            instance_lst.append({'property': pro, 'instances': instance_ids})
        lst.append({'id': key, 'groups': instance_lst})
    return JsonResponse(lst, safe=False)


@login_required
def items_locations_as_json(request):
    items_locations = ItemLocation.objects.all()
    items_locations_data = ItemLocationSerializer(items_locations, many=True).data
    return JsonResponse(items_locations_data, safe=False)


# class ItemLocationAPI(generics.ListCreateAPIView):
#     queryset = ItemLocation.objects.all()
#     serializer_class = ItemLocationSerializer
#     permission_classes = (permissions.IsAuthenticatedOrReadOnly,)


def create_item_location(request, id=None):
    if id:
        il = get_object_or_404(ItemLocation, id=id)
        scenario = 'Edit'
    else:
        il = ItemLocation()
        scenario = 'Create'
    if request.POST:
        form = ItemLocationForm(request.POST, instance=il)
        if form.is_valid():
            i_loc = form.save()
        if request.is_ajax():
            return render(request, 'callback.html', {'obj': ItemLocationSerializer(i_loc).data})
        return redirect('/')
    form = ItemLocationForm(instance=il)
    if request.is_ajax():
        base_template = 'modal.html'
    else:
        base_template = 'inventory_base.html'
    return render(request, 'item_form.html', {
        'scenario': scenario,
        'form': form,
        'base_template': base_template,
    })


@group_required('Store Keeper', 'Chief')
def list_categories(request):
    categories = Category.objects.filter()
    return render(request, 'list_inventory_categories.html', {'categories': categories})


@group_required('Store Keeper', 'Chief')
def create_category(request):
    category = Category()
    if request.POST:
        form = CategoryForm(data=request.POST)
        if form.is_valid():
            category = form.save(commit=False)
            category.save()
            return redirect('/inventory/categories/')
    else:
        form = CategoryForm(instance=category)
    if request.is_ajax():
        base_template = 'modal.html'
    else:
        base_template = 'inventory_base.html'
    return render(request, 'inventory_category_create_form.html', {
        'form': form,
        'base_template': base_template,
    })


@group_required('Store Keeper', 'Chief')
def update_category(request, id):
    category = get_object_or_404(Category, id=id)
    if request.POST:
        form = CategoryForm(data=request.POST, instance=category)
        if form.is_valid():
            category = form.save(commit=False)
            category.save()
            return redirect('/inventory/categories/')
    else:
        form = CategoryForm(instance=category)
    if request.is_ajax():
        base_template = 'modal.html'
    else:
        base_template = 'inventory_base.html'
    return render(request, 'inventory_category_update_form.html', {
        'form': form,
        'base_template': base_template
    })


@group_required('Store Keeper', 'Chief')
def delete_category(request, id):
    category = get_object_or_404(Category, id=id)
    category.delete()
    return redirect('/inventory/categories/')


@login_required
def demand_form(request, id=None):
    if id:
        obj = get_object_or_404(Demand, id=id)
        scenario = 'Update'
    else:
        obj = Demand(demandee=request.user)
        scenario = 'Create'
    form = DemandForm(instance=obj)
    object_data = DemandSerializer(obj).data
    return render(request, 'demand_form.html',
                  {'form': form, 'data': object_data, 'scenario': scenario})


@login_required
def save_demand(request):
    params = json.loads(request.body)
    dct = {'rows': {}}
    if params.get('release_no') == '':
        params['release_no'] = None
    object_values = {'release_no': params.get('release_no'), 'fiscal_year': FiscalYear.get(app_setting.fiscal_year),
                     'date': params.get('date'), 'purpose': params.get('purpose'), 'status': 'Requested'}
    if params.get('id'):
        obj = Demand.objects.get(id=params.get('id'))
    else:
        obj = Demand()
        object_values['demandee_id'] = params.get('demandee')
    # try:
    if True:
        obj = save_model(obj, object_values)
        dct['id'] = obj.id
        model = DemandRow
        for index, row in enumerate(params.get('table_view').get('rows')):
            invalid_check = invalid(row, ['item_id', 'quantity', 'unit', 'purpose'])
            if invalid_check:
                dct['error_message'] = 'These feilds must be filled: ' + ', '.join(invalid_check)
                return JsonResponse(dct)
            else:
                # if row.get('release_quantity') == '':
                # row['release_quantity'] = 1
                values = {'sn': index + 1, 'item_id': row.get('item_id'),
                          'specification': row.get('specification'),
                          'quantity': row.get('quantity'), 'unit': row.get('unit'), 'remarks': row.get('remarks'),
                          'purpose': row.get('purpose'), 'demand': obj}

                submodel, created = model.objects.get_or_create(id=row.get('id'), defaults=values)

                submodel.releases.all().delete()
                for release in row['release_vms']:
                    for instance in release['instances']:
                        instance_model = ItemInstance.objects.get(id=instance)
                        instance_model.location_id = release['location_id']
                        instance_model.save()
                        rel = Release(item_instance=instance_model, demand_row=submodel)
                        rel.save()

            # set_transactions(submodel, request.POST.get('date'),
            #                  ['dr', bank_account, row.get('amount')],
            #                  ['cr', benefactor, row.get('amount')],
            # )
            if not created:
                submodel = save_model(submodel, values)
            dct['rows'][index] = submodel.id
        delete_rows(params.get('table_view').get('deleted_rows'), model)
    # except Exception as e:
    #     if hasattr(e, 'messages'):
    #         dct['error_message'] = '; '.join(e.messages)
    #     elif str(e) != '':
    #         dct['error_message'] = str(e)
    #     else:
    #         dct['error_message'] = 'Error in form data!'
    return JsonResponse(dct)


@group_required('Store Keeper', 'Chief')
def purchase_order(request, id=None):
    if id:
        obj = get_object_or_404(PurchaseOrder, id=id)
        scenario = 'Update'
    else:
        obj = PurchaseOrder()
        scenario = 'Create'
    form = PurchaseOrderForm(instance=obj)
    object_data = PurchaseOrderSerializer(obj).data
    return render(request, 'purchase_order.html',
                  {'form': form, 'data': object_data, 'scenario': scenario})


@group_required('Store Keeper', 'Chief')
def save_purchase_order(request):
    params = json.loads(request.body)
    dct = {'rows': {}}
    object_values = {'order_no': empty_to_none(params.get('order_no')),
                     'fiscal_year': FiscalYear.get(app_setting.fiscal_year),
                     'date': params.get('date'), 'party_id': params.get('party'),
                     'due_days': params.get('due_days')}
    if params.get('id'):
        obj = PurchaseOrder.objects.get(id=params.get('id'))
    else:
        obj = PurchaseOrder()
    try:
        obj = save_model(obj, object_values)
        dct['id'] = obj.id
        model = PurchaseOrderRow
        for index, row in enumerate(params.get('table_view').get('rows')):
            if invalid(row, ['quantity', 'unit', 'rate', 'item_id']):
                continue
            if row.get('budget_title_no') == '':
                row['budget_title_no'] = None
            values = {'sn': index + 1, 'item_id': row.get('item_id'),
                      'specification': row.get('specification'), 'rate': row.get('rate'),
                      'quantity': row.get('quantity'), 'unit': row.get('unit'), 'vattable': row.get('vattable'),
                      'budget_title_no': row.get('budget_title_no'), 'remarks': row.get('remarks'),
                      'purchase_order': obj}

            submodel, created = model.objects.get_or_create(id=row.get('id'), defaults=values)
            # set_transactions(submodel, request.POST.get('date'),
            # ['dr', bank_account, row.get('amount')],
            #                  ['cr', benefactor, row.get('amount')],
            # )
            if not created:
                submodel = save_model(submodel, values)
            dct['rows'][index] = submodel.id
        delete_rows(params.get('table_view').get('deleted_rows'), model)
    except Exception as e:
        if hasattr(e, 'messages'):
            dct['error_message'] = '; '.join(e.messages)
        elif str(e) != '':
            dct['error_message'] = str(e)
        else:
            dct['error_message'] = 'Error in form data!'
    return JsonResponse(dct)


@group_required('Store Keeper', 'Chief')
def list_purchase_orders(request):
    objects = PurchaseOrder.objects.all()
    return render(request, 'list_purchase_orders.html', {'objects': objects})


@group_required('Store Keeper', 'Chief')
def delete_purchase_order(request, id):
    obj = get_object_or_404(PurchaseOrder, id=id)
    obj.delete()
    return redirect(reverse('list_purchase_orders'))


@group_required('Store Keeper', 'Chief')
def list_inventory_accounts(request):
    objects = InventoryAccount.objects.all()
    return render(request, 'list_inventory_accounts.html', {'objects': objects})


@group_required('Store Keeper', 'Chief')
def list_consumable_accounts(request):
    objects = InventoryAccount.objects.filter(item__type='consumable')
    return render(request, 'list_consumable_inventory_accounts.html', {'objects': objects})


@group_required('Store Keeper', 'Chief')
def list_non_consumable_accounts(request):
    objects = InventoryAccount.objects.filter(item__type='non-consumable')
    return render(request, 'list_non_consumable_inventory_accounts.html', {'objects': objects})


@group_required('Store Keeper', 'Chief')
def view_inventory_account(request, id):
    obj = get_object_or_404(InventoryAccount, id=id)
    journal_entries = JournalEntry.objects.filter(transactions__account_id=obj.id).order_by('id', 'date') \
        .prefetch_related('transactions', 'content_type', 'transactions__account').select_related()
    data = InventoryAccountRowSerializer(journal_entries).data
    return render(request, 'view_inventory_account.html', {'obj': obj, 'entries': journal_entries, 'data': data})


@group_required('Store Keeper', 'Chief')
def handover_incoming(request, id=None):
    if id:
        obj = get_object_or_404(Handover, id=id)
        scenario = 'Update'
    else:
        obj = Handover(type='Incoming')
        scenario = 'Create'
    form = HandoverForm(instance=obj)
    object_data = HandoverSerializer(obj).data
    return render(request, 'handover.html',
                  {'form': form, 'data': object_data, 'scenario': scenario})


@group_required('Store Keeper', 'Chief')
def handover_outgoing(request, id=None):
    if id:
        obj = get_object_or_404(Handover, id=id)
        scenario = 'Update'
    else:
        obj = Handover(type='Outgoing')
        scenario = 'Create'
    form = HandoverForm(instance=obj)
    object_data = HandoverSerializer(obj).data
    return render(request, 'handover.html',
                  {'form': form, 'data': object_data, 'scenario': scenario})


@group_required('Store Keeper', 'Chief')
def save_handover(request):
    params = json.loads(request.body)
    dct = {'rows': {}}
    object_values = {'addressee': params.get('addressee'), 'fiscal_year': FiscalYear.get(app_setting.fiscal_year),
                     'date': params.get('date'), 'office': params.get('office'), 'type': params.get('type'),
                     'designation': params.get('designation'), 'voucher_no': empty_to_none(params.get('voucher_no')),
                     'due_days': params.get('due_days'), 'handed_to': params.get('handed_to')}
    if params.get('id'):
        obj = Handover.objects.get(id=params.get('id'))
    else:
        obj = Handover()
    try:
        obj = save_model(obj, object_values)
    except Exception as e:
        if hasattr(e, 'messages'):
            dct['error_message'] = '; '.join(e.messages)
        elif str(e) != '':
            dct['error_message'] = str(e)
        else:
            dct['error_message'] = 'Error in form data!'
    dct['id'] = obj.id
    model = HandoverRow
    for index, row in enumerate(params.get('table_view').get('rows')):
        if invalid(row, ['quantity', 'unit', 'item_id', 'total_amount']):
            continue
        values = {'sn': index + 1, 'item_id': row.get('item_id'),
                  'specification': row.get('specification'),
                  'quantity': row.get('quantity'), 'unit': row.get('unit'), 'received_date': row.get('received_date'),
                  'total_amount': row.get('total_amount'), 'condition': row.get('condition'),
                  'handover': obj}
        submodel, created = model.objects.get_or_create(id=row.get('id'), defaults=values)
        if not created:
            submodel = save_model(submodel, values)
        dct['rows'][index] = submodel.id
    delete_rows(params.get('table_view').get('deleted_rows'), model)
    return JsonResponse(dct)


@group_required('Store Keeper', 'Chief')
def list_incoming_handovers(request):
    objects = Handover.objects.filter(type='Incoming')
    return render(request, 'list_incoming_handovers.html', {'objects': objects})


@group_required('Store Keeper', 'Chief')
def list_outgoing_handovers(request):
    objects = Handover.objects.filter(type='Outgoing')
    return render(request, 'list_outgoing_handovers.html', {'objects': objects})


@group_required('Store Keeper', 'Chief')
def handover_entry_report(request, id=None):
    source = get_object_or_404(Handover, id=id, type='Incoming')
    if source.get_entry_report():
        report = source.get_entry_report()
        object_data = EntryReportSerializer(report).data
    else:
        report = EntryReport()
        object_data = EntryReportSerializer(report).data
        report.fiscal_year = source.fiscal_year
        report.source = source
        all_rows = []
        for r in source.rows.all():
            row = EntryReportRow()
            row.sn = r.sn
            row.item = r.item
            row.specification = r.specification
            row.quantity = r.quantity
            row.unit = r.unit
            row.rate = r.total_amount / r.quantity
            row.remarks = r.condition
            row_data = EntryReportRowSerializer(row).data
            all_rows.append(row_data)
        object_data.update({'rows': all_rows})
    form = EntryReportForm(instance=report)
    object_data['type'] = 'handover'
    object_data['source_id'] = source.id
    return render(request, 'entry_report.html',
                  {'form': form, 'data': object_data})


@group_required('Store Keeper', 'Chief')
def purchase_entry_report(request, id=None):
    source = get_object_or_404(PurchaseOrder, id=id)
    if source.get_entry_report():
        report = source.get_entry_report()
        object_data = EntryReportSerializer(report).data
    else:
        report = EntryReport()
        object_data = EntryReportSerializer(report).data
        report.fiscal_year = source.fiscal_year
        report.source = source
        all_rows = []
        for r in source.rows.all():
            row = EntryReportRow()
            row.sn = r.sn
            row.item = r.item
            row.specification = r.specification
            row.quantity = r.quantity
            row.unit = r.unit
            row.rate = r.rate
            row.remarks = r.remarks
            row_data = EntryReportRowSerializer(row).data
            all_rows.append(row_data)
        object_data.update({'rows': all_rows})
    form = EntryReportForm(instance=report)
    object_data['type'] = 'purchase'
    object_data['source_id'] = source.id
    return render(request, 'entry_report.html',
                  {'form': form, 'data': object_data})


@group_required('Store Keeper', 'Chief')
def save_entry_report(request):
    params = json.loads(request.body)
    dct = {'rows': {}}
    if params.get('type') == 'handover':
        source = Handover.objects.get(id=params.get('source_id'))
    else:
        source = PurchaseOrder.objects.get(id=params.get('source_id'))
    object_values = {'entry_report_no': empty_to_none(params.get('entry_report_no')),
                     'fiscal_year': FiscalYear.get(app_setting.fiscal_year),
                     'source': source}
    if params.get('id'):
        obj = EntryReport.objects.get(id=params.get('id'))
    else:
        obj = EntryReport()
    try:
        obj = save_model(obj, object_values)
    except Exception as e:
        if hasattr(e, 'messages'):
            dct['error_message'] = '; '.join(e.messages)
        elif str(e) != '':
            dct['error_message'] = str(e)
        else:
            dct['error_message'] = 'Error in form data!'
    dct['id'] = obj.id
    model = EntryReportRow
    for index, row in enumerate(params.get('table_view').get('rows')):
        if invalid(row, ['quantity', 'unit', 'item_id', 'rate']):
            continue
        if row.get('other_expenses') == '':
            other_expenses = 0
        else:
            other_expenses = row.get('other_expenses')
        values = {'sn': index + 1, 'item_id': row.get('item_id'),
                  'specification': row.get('specification'),
                  'quantity': row.get('quantity'), 'unit': row.get('unit'), 'rate': row.get('rate'),
                  'remarks': row.get('remarks'), 'other_expenses': other_expenses,
                  'entry_report': obj}
        submodel, created = model.objects.get_or_create(id=row.get('id'), defaults=values)
        if not created:
            submodel = save_model(submodel, values)
        dct['rows'][index] = submodel.id
        set_transactions(submodel, obj.source.date,
                         ['dr', submodel.item.account, submodel.quantity],
                         )
        for i in range(0, int(row.get('quantity'))):
            item_instance = ItemInstance()
            item_instance.item = Item.objects.get(id=int(row.get('item_id')))
            item_instance.item_rate = row.get('rate')
            item_instance.location = ItemLocation.objects.get(name='Store')
            item_instance.source = submodel
            item_instance.other_properties = item_instance.item.other_properties
            item_instance.save()
    delete_rows(params.get('table_view').get('deleted_rows'), model)
    return JsonResponse(dct)


@group_required('Store Keeper', 'Chief')
def list_entry_reports(request):
    objects = EntryReport.objects.all()
    return render(request, 'list_entry_reports.html', {'objects': objects})


@group_required('Store Keeper', 'Chief')
def list_handover_entry_reports(request):
    objects = EntryReport.objects.filter(source_content_type__model='handover')
    return render(request, 'list_entry_reports.html', {'objects': objects})


@group_required('Store Keeper', 'Chief')
def list_purchase_entry_reports(request):
    objects = EntryReport.objects.filter(source_content_type__model='purchaseorder')
    return render(request, 'list_entry_reports.html', {'objects': objects})


@group_required('Store Keeper', 'Chief')
def delete_entry_report(request, id):
    obj = get_object_or_404(EntryReport, id=id)
    obj.delete()
    return redirect(reverse('list_entry_reports'))


@group_required('Store Keeper', 'Chief')
def approve_demand(request):
    params = json.loads(request.body)
    dct = {'rows': {}}
    if params.get('id'):
        row = DemandRow.objects.get(id=params.get('id'))
    else:
        dct['error_message'] = 'Row needs to be saved before being approved!'
        return JsonResponse(dct)
    invalid_check = invalid(params, ['item_id', 'quantity', 'unit', 'release_quantity', 'purpose', 'location'])
    if invalid_check:
        dct['error_message'] = 'Fill out following fields: ' + ', '.join(invalid_check)
        return JsonResponse(dct)
    else:
        if row.item.account.current_balance < float(params.get('release_quantity')):
            dct['error_message'] = 'We dont have this item in stock. Purchase Item'
        values = {'item_id': params.get('item_id'),
                  'specification': params.get('specification'),
                  'quantity': params.get('quantity'), 'unit': params.get('unit'),
                  'release_quantity': params.get('release_quantity'), 'remarks': params.get('remarks'),
                  'purpose': params.get('purpose'), 'location': ItemLocation.objects.get(id=params.get('location'))}
        row = save_model(row, values)
    row.status = 'Approved'
    row.save()
    return JsonResponse(dct)


@group_required('Store Keeper', 'Chief')
def disapprove_demand(request):
    params = json.loads(request.body)
    dct = {}
    if params.get('id'):
        row = DemandRow.objects.get(id=params.get('id'))
    else:
        dct['error_message'] = 'Voucher needs to be saved before being disapproved!'
        return JsonResponse(dct)
    row.status = 'Requested'
    row.save()
    return JsonResponse(dct)


@group_required('Store Keeper', 'Chief')
def fulfill_demand(request):
    params = json.loads(request.body)
    dct = {}
    if params.get('id'):
        row = DemandRow.objects.get(id=params.get('id'))
    else:
        dct['error_message'] = 'Row needs to be saved before being fulfilled!'
        return JsonResponse(dct)
    if params['status'] == 'Requested':
        dct['error_message'] = 'Row needs to be approved before being fulfilled!'
        return JsonResponse(dct)
    set_transactions(row, row.demand.date,
                     ['cr', row.item.account, row.release_quantity],
                     )

    # Search items in the stock
    items = ItemInstance.objects.filter(item_id=row.item_id, location_id=STORE_LOCATION_ID)
    release_quantity = int(row.release_quantity)
    for item, i in zip(items, range(0, release_quantity)):
        item.location = row.location
        item.save()

    row.status = 'Fulfilled'
    row.save()
    return JsonResponse(dct)


@group_required('Store Keeper', 'Chief')
def unfulfill_demand(request):
    params = json.loads(request.body)
    dct = {}
    if params.get('id'):
        row = DemandRow.objects.get(id=params.get('id'))
    else:
        dct['error_message'] = 'Row needs to be saved before being unfulfilled!'
        return JsonResponse(dct)
    if params['status'] != 'Fulfilled':
        dct['error_message'] = 'Row needs to be fulfilled before being unfulfilled!'
        return JsonResponse(dct)
    journal_entry = JournalEntry.get_for(row)
    journal_entry.delete()

    items = ItemInstance.objects.filter(item_id=row.item_id, location_id=int(params.get('location')))
    for item in items:
        item.location = ItemLocation.objects.get(name='Store')
        item.save()

    row.status = 'Approved'
    row.save()
    return JsonResponse(dct)


@group_required('Store Keeper', 'Chief')
def save_account(request):
    params = json.loads(request.body)
    dct = {'rows': {}}
    for index, row in enumerate(params.get('table_vm').get('rows')):
        entry = JournalEntry.objects.get(id=row.get('id'))
        try:
            account_row = entry.account_row
        except:
            account_row = InventoryAccountRow(journal_entry=entry)
        if row.get('expense_total_cost_price') == '':
            row['expense_total_cost_price'] = None
        if row.get('remaining_total_cost_price') == '':
            row['remaining_total_cost_price'] = None
        values = {'country_of_production_or_company_name': row.get('country_or_company'), 'size': row.get('size'),
                  'expected_life': row.get('expected_life'), 'source': row.get('source'), 'remarks': row.get('remarks'),
                  'expense_total_cost_price': row.get('expense_total_cost_price'),
                  'remaining_total_cost_price': row.get('remaining_total_cost_price')}
        account_row = save_model(account_row, values)
    return JsonResponse(dct)


def index(request):
    return render(request, 'inventory_index.html')
